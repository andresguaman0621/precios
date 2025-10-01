from flask import Flask, render_template, request, jsonify, send_file
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

app = Flask(__name__)

# Archivos para almacenar los datos
DATA_FILE = 'productos_data.json'
MARISCOS_DATA_FILE = 'mariscos_data.json'

def load_data(file_path=DATA_FILE):
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_data(data, file_path=DATA_FILE):
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

@app.route('/')
def index():
    productos = load_data()
    categorias = list(set(p['categoria'] for p in productos if p.get('categoria')))
    return render_template('index.html', productos=productos, categorias=categorias)

@app.route('/api/productos')
def api_productos():
    productos = load_data()
    categoria = request.args.get('categoria', '')
    codigo = request.args.get('codigo', '')
    nombre = request.args.get('nombre', '')
    
    # Aplicar filtros
    if categoria:
        productos = [p for p in productos if p.get('categoria', '').upper() == categoria.upper()]
    if codigo:
        productos = [p for p in productos if str(p.get('codigo', '')).startswith(codigo)]
    if nombre:
        productos = [p for p in productos if nombre.upper() in p.get('nombre_completo', '').upper()]
    
    return jsonify(productos)

@app.route('/api/guardar_precio', methods=['POST'])
def guardar_precio():
    data = request.json
    codigo = data.get('codigo')
    precio_actual = data.get('precio_actual')

    productos = load_data()
    for producto in productos:
        if producto['codigo'] == codigo:
            producto['precio_actual'] = precio_actual
            break

    save_data(productos)
    return jsonify({'success': True})

@app.route('/api/guardar_todos', methods=['POST'])
def guardar_todos():
    data = request.json
    productos = load_data()

    for item in data:
        codigo = item.get('codigo')
        precio_actual = item.get('precio_actual')

        for producto in productos:
            if producto['codigo'] == codigo:
                producto['precio_actual'] = precio_actual
                break

    save_data(productos)
    return jsonify({'success': True})

@app.route('/api/nueva_toma', methods=['POST'])
def nueva_toma():
    productos = load_data()

    for producto in productos:
        producto['precio_anterior'] = producto.get('precio_actual', '')
        producto['precio_actual'] = ''

    save_data(productos)
    return jsonify({'success': True})

@app.route('/api/export_excel')
def export_excel():
    productos = load_data()
    categoria = request.args.get('categoria', '')
    codigo = request.args.get('codigo', '')
    nombre = request.args.get('nombre', '')

    # Aplicar los mismos filtros que en /api/productos
    if categoria:
        productos = [p for p in productos if p.get('categoria', '').upper() == categoria.upper()]
    if codigo:
        productos = [p for p in productos if str(p.get('codigo', '')).startswith(codigo)]
    if nombre:
        productos = [p for p in productos if nombre.upper() in p.get('nombre_completo', '').upper()]

    # Crear libro de trabajo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Precios"

    # Definir encabezados
    headers = ['Código', 'Estado', 'Categoría', 'Producto', 'Tamaño', 'Presentación', 'Peso (Kg)', 'Precio Anterior', 'Precio Actual']

    # Agregar encabezados con formato
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos
    for row, producto in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=producto.get('codigo', ''))
        ws.cell(row=row, column=2, value=producto.get('estado', ''))
        ws.cell(row=row, column=3, value=producto.get('categoria', ''))
        ws.cell(row=row, column=4, value=producto.get('nombre_completo', ''))
        ws.cell(row=row, column=5, value=producto.get('tamaño', ''))
        ws.cell(row=row, column=6, value=producto.get('presentacion', ''))
        ws.cell(row=row, column=7, value=producto.get('peso_kg', ''))

        # Convertir precios a números si tienen valor
        precio_anterior = producto.get('precio_anterior', '')
        precio_actual = producto.get('precio_actual', '')

        try:
            precio_anterior = float(precio_anterior) if precio_anterior else ''
        except (ValueError, TypeError):
            precio_anterior = precio_anterior

        try:
            precio_actual = float(precio_actual) if precio_actual else ''
        except (ValueError, TypeError):
            precio_actual = precio_actual

        ws.cell(row=row, column=8, value=precio_anterior)
        ws.cell(row=row, column=9, value=precio_actual)

    # Ajustar ancho de columnas
    column_widths = [10, 8, 12, 40, 12, 15, 12, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Crear archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Crear nombre de archivo con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reporte_precios_{timestamp}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# RUTAS PARA MARISCOS
@app.route('/mariscos')
def mariscos():
    productos = load_data(MARISCOS_DATA_FILE)
    return render_template('mariscos.html', productos=productos)

@app.route('/api/mariscos')
def api_mariscos():
    productos = load_data(MARISCOS_DATA_FILE)
    nombre = request.args.get('nombre', '')

    # Aplicar filtro por nombre
    if nombre:
        productos = [p for p in productos if nombre.upper() in p.get('producto', '').upper()]

    return jsonify(productos)

@app.route('/api/mariscos/guardar_precio', methods=['POST'])
def guardar_precio_mariscos():
    data = request.json
    producto_id = data.get('id')
    precio_actual = data.get('precio_actual')

    productos = load_data(MARISCOS_DATA_FILE)
    for producto in productos:
        if producto['id'] == producto_id:
            producto['precio_actual'] = precio_actual
            break

    save_data(productos, MARISCOS_DATA_FILE)
    return jsonify({'success': True})

@app.route('/api/mariscos/guardar_todos', methods=['POST'])
def guardar_todos_mariscos():
    data = request.json
    productos = load_data(MARISCOS_DATA_FILE)

    for item in data:
        producto_id = item.get('id')
        precio_actual = item.get('precio_actual')

        for producto in productos:
            if producto['id'] == producto_id:
                producto['precio_actual'] = precio_actual
                break

    save_data(productos, MARISCOS_DATA_FILE)
    return jsonify({'success': True})

@app.route('/api/mariscos/nueva_toma', methods=['POST'])
def nueva_toma_mariscos():
    productos = load_data(MARISCOS_DATA_FILE)

    for producto in productos:
        producto['precio_anterior'] = producto.get('precio_actual', '')
        producto['precio_actual'] = ''

    save_data(productos, MARISCOS_DATA_FILE)
    return jsonify({'success': True})

@app.route('/api/mariscos/export_excel')
def export_excel_mariscos():
    productos = load_data(MARISCOS_DATA_FILE)
    nombre = request.args.get('nombre', '')

    # Aplicar filtro
    if nombre:
        productos = [p for p in productos if nombre.upper() in p.get('producto', '').upper()]

    # Crear libro de trabajo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Precios Mariscos"

    # Definir encabezados
    headers = ['ID', 'Producto', 'Peso', 'Precio Anterior', 'Precio Actual']

    # Agregar encabezados con formato
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos
    for row, producto in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=producto.get('id', ''))
        ws.cell(row=row, column=2, value=producto.get('producto', ''))
        ws.cell(row=row, column=3, value=producto.get('peso', ''))

        # Convertir precios a números si tienen valor
        precio_anterior = producto.get('precio_anterior', '')
        precio_actual = producto.get('precio_actual', '')

        try:
            precio_anterior = float(precio_anterior) if precio_anterior else ''
        except (ValueError, TypeError):
            precio_anterior = precio_anterior

        try:
            precio_actual = float(precio_actual) if precio_actual else ''
        except (ValueError, TypeError):
            precio_actual = precio_actual

        ws.cell(row=row, column=4, value=precio_anterior)
        ws.cell(row=row, column=5, value=precio_actual)

    # Ajustar ancho de columnas
    column_widths = [8, 40, 15, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Crear archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Crear nombre de archivo con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reporte_mariscos_{timestamp}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)